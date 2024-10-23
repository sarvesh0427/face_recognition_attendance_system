import face_recognition
import cv2
import numpy as np
from openpyxl import load_workbook
from datetime import datetime
import time
import pygame

# This project is a face recognition attendance system using OpenCV and an Excel file as a database.
# It captures student images and stores them in a photodatabase Excel file with the corresponding student names.
# The system then marks attendance by writing 'P' to the student names into an AttendanceSheet if student is present on that day.
# The attendance will be marked for 15 days, and the real date will be recorded for each entry.

# Load the Excel workbooks
photo_db_wb = load_workbook("photodatabase.xlsx")
photo_db_ws = photo_db_wb.active
attendance_wb = load_workbook("attendancesheet.xlsx")
attendance_ws = attendance_wb.active

# Set the timer for 5 minutes (300 seconds)
TIMER = 1 * 60
start_time = time.time()

# Initialize the sound system for attendance confirmation
pygame.mixer.init()
sound1 = pygame.mixer.Sound("present.mp3")

# Set up the current date and find the corresponding column in the attendance sheet
now1 = datetime.now()
current_date = now1.strftime("%Y-%m-%d")

# Find the correct column for today's attendance in 'attendancesheet.xlsx'
column_found = False
for col in range(3, 17):  # Assuming C2 to P2 for attendance
    cell = attendance_ws.cell(row=2, column=col)
    if cell.value is None:  # Find the first empty cell
        cell.value = current_date  # Set the current date
        attendance_wb.save('attendancesheet.xlsx')
        attendance_column = col  # Store the column index for attendance
        column_found = True
        break

if not column_found:
    print("No available column for today's attendance.")
    exit()

# Load known face encodings and student names from 'photo_database.xlsx'
known_face_encodings = []
known_face_names = []

for row in range(2, photo_db_ws.max_row + 1):  # Assuming data starts from row 2
    image_path = photo_db_ws.cell(row=row, column=3).value  # Image path in column A
    student_name = photo_db_ws.cell(row=row, column=2).value  # Student name in column B

    # Load the image and encode the face
    try:
        if image_path:
            print(f"Loading image: {image_path}")
            student_image = face_recognition.load_image_file(image_path)
            student_face_encoding = face_recognition.face_encodings(student_image)
            if student_face_encoding:
                known_face_encodings.append(student_face_encoding[0])
                known_face_names.append(student_name)
            else:
                print(f"No face detected in the image at {image_path}")
        else:
            print(f"Invalid image path at row {row}")
    except Exception as e:
        print(f"Error loading or encoding image at {image_path}: {e}")

# Initialize video capture (webcam)
video_capture = cv2.VideoCapture(0)

# Initialize variables for face detection and processing
face_locations = []
face_encodings = []
face_names = []
process_this_frame = True

# Track who has already been marked present
attendance_marked = {name: False for name in known_face_names}

while True:
    # Capture a single frame from the webcam
    ret, frame = video_capture.read()

    # Process every other frame for efficiency
    if process_this_frame:
        # **Use a larger frame size for more accurate recognition**
        small_frame = cv2.resize(frame, (0, 0), fx=0.55, fy=0.55)  # Adjust size if needed
        rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

        # Face detection and recognition
        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

        face_names = []
        for face_encoding in face_encodings:
            # Adjust the tolerance to make matching less strict
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding, tolerance=0.5)
            name = "Unknown"

            # Find the best match, only if we have valid face encodings
            if len(known_face_encodings) > 0:
                face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
                if len(face_distances) > 0:
                    best_match_index = np.argmin(face_distances)
                    if matches[best_match_index]:
                        name = known_face_names[best_match_index]

            face_names.append(name)

    process_this_frame = not process_this_frame

    # Display the reverse countdown timer in minutes and seconds format
    elapsed_time = int(time.time() - start_time)
    remaining_time = TIMER - elapsed_time
    minutes = remaining_time // 60
    seconds = remaining_time % 60
    timer_text = f"{minutes:02}:{seconds:02}"

    # Draw a frame with the heading and timer
    font = cv2.FONT_HERSHEY_SIMPLEX
    heading_text = "Attendance Remaining Time"

    # Create a semi-transparent rectangle for the timer display
    overlay = frame.copy()
    cv2.rectangle(overlay, (10, 10), (250, 100), (0, 0, 0), cv2.FILLED)
    frame = cv2.addWeighted(overlay, 0.6, frame, 0.4, 0)

    # Draw the heading and timer text
    cv2.putText(frame, heading_text, (20, 40), font, 0.5, (255, 255, 255), 2, cv2.LINE_AA)
    cv2.putText(frame, timer_text, (20, 80), font, 1.0, (0, 255, 0), 2, cv2.LINE_AA)

    # Mark attendance based on recognized face and play sound once
    for (top, right, bottom, left), name in zip(face_locations, face_names):
        top *= 2  # Adjust the scaling factor based on the new frame size
        right *= 2
        bottom *= 2
        left *= 2

        # Set the frame color to green if the face is recognized and present, otherwise red
        if name != "Unknown" and attendance_marked[name]:
            frame_color = (0, 255, 0)  # Green for recognized faces marked present
        else:
            frame_color = (0, 0, 255)  # Red for unrecognized faces or not marked present

        # Draw a box around the face and label it
        cv2.rectangle(frame, (left, top), (right, bottom), frame_color, 2)
        cv2.rectangle(frame, (left, bottom - 35), (right, bottom), frame_color, cv2.FILLED)
        cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)

        # Update attendance sheet based on recognized face, and only play sound once per person
        if name != "Unknown" and not attendance_marked[name]:
            # Find the correct row in attendancesheet.xlsx for this student
            for row in range(3, attendance_ws.max_row+1):  # Assuming student names are in rows 3 to 13
                if attendance_ws.cell(row=row, column=2).value == name:
                    attendance_ws.cell(row=row, column=attendance_column).value = 'P'
                    attendance_wb.save('attendancesheet.xlsx')
                    sound1.play()  # Play sound
                    attendance_marked[name] = True  # Mark attendance as done
                    break

    # If no face is detected, mark everyone as "A" (absent)
    if len(face_names) == 0:
        for row in range(3, attendance_ws.max_row+1):
            if attendance_ws.cell(row=row, column=attendance_column).value != 'P':
                attendance_ws.cell(row=row, column=attendance_column).value = 'A'
                attendance_wb.save('attendancesheet.xlsx')

    # End the program when the timer runs out
    if remaining_time <= 0:
        print("Time's up! Quitting the program.")
        break

    # Show the video frame
    cv2.imshow('Attendance System', frame)

    # Exit the loop when 'q' is pressed
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release the webcam and close windows
video_capture.release()
cv2.destroyAllWindows()
